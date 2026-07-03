# ─────────────────────────────────────────────────────────────
# KCIM 일가정양립 지원 가이드 v2.0  (경영관리본부 담당자용)
# 기준: 2025-02-23 시행 개정법 / KCIM 2026 복지제도 / 최종 검토 2026-07-04
# 필요 패키지: streamlit, pandas, openpyxl, openai  (requirements.txt 참조)
# 실행: streamlit run app.py
# 챗봇 사용 시 .streamlit/secrets.toml 에 OPENAI_API_KEY 필요
# ─────────────────────────────────────────────────────────────
import json
import math
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st

st.set_page_config(page_title="KCIM 일가정양립 지원 가이드", page_icon="🤝", layout="wide")

VERSION_BASIS = "법정 제도: 2025-02-23 시행 개정법 기준 · 회사 복지: 2026 복지제도 · 최종 검토 2026-07-04"
CASES_FILE = Path(__file__).parent / "kcim_cases.json"

# ─────────────────────────────────────────────────────────────
# CSS  (폰트 최소 12px / 색상 4계열: 파랑=법정·주요, 초록=회사·완료, 주황=주의, 빨강=경고)
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
*, html, body, [class*="css"] { font-family: 'Pretendard', -apple-system, sans-serif !important; box-sizing: border-box; }
.block-container { padding: 0 1.2rem 2rem !important; max-width: 1240px !important; }
.stApp { background: #f5f7fa; }
section[data-testid="stSidebar"], header[data-testid="stHeader"], div[data-testid="stToolbar"] { display: none !important; }

.top-header { background: #16395c; padding: 0.9rem 1.4rem; border-radius: 0 0 12px 12px;
  display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px; margin-bottom: 0.9rem; }
.top-title { font-size: 1.15rem; font-weight: 800; color: #fff; letter-spacing: -0.3px; }
.top-sub { font-size: 0.75rem; color: rgba(255,255,255,0.75); margin-top: 2px; }
.ver-badge { background: #e8f1fb; color: #16395c; padding: 4px 12px; border-radius: 16px; font-size: 0.75rem; font-weight: 700; }

.sec-title { font-size: 0.8rem; font-weight: 800; color: #475569; margin: 0.9rem 0 0.4rem; }

.scn-card { background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 0.9rem; height: 100%; }
.scn-name { font-size: 0.92rem; font-weight: 800; color: #1e293b; margin: 4px 0 2px; }
.scn-desc { font-size: 0.75rem; color: #64748b; line-height: 1.5; }

.tl-wrap { background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 0.7rem 1rem; margin-bottom: 0.8rem;
  display: flex; align-items: center; gap: 4px; overflow-x: auto; }
.tl-dot { width: 26px; height: 26px; min-width: 26px; border-radius: 50%; display: flex; align-items: center;
  justify-content: center; font-size: 0.75rem; font-weight: 800; border: 2px solid #cbd5e1; color: #94a3b8; background: #f8fafc; }
.tl-dot.done { background: #dcfce7; border-color: #22c55e; color: #15803d; }
.tl-dot.now { background: #2563eb; border-color: #2563eb; color: #fff; width: 30px; height: 30px; min-width: 30px; }
.tl-line { flex: 1; height: 2px; background: #e2e8f0; min-width: 10px; }
.tl-label { font-size: 0.75rem; font-weight: 700; color: #334155; margin-left: 10px; white-space: nowrap; }

.step-head { border-radius: 12px; padding: 1rem 1.3rem; margin-bottom: 0.8rem; color: #fff; background: #16395c; }
.step-head .num { font-size: 0.72rem; font-weight: 700; background: rgba(255,255,255,0.2); border-radius: 14px; padding: 2px 10px; display: inline-block; }
.step-head .ttl { font-size: 1.25rem; font-weight: 900; margin-top: 4px; }
.step-head .meta { font-size: 0.76rem; opacity: 0.85; margin-top: 4px; }

.law-card { background: #f0f6fd; border: 1px solid #bcd7f2; border-radius: 12px; padding: 0.9rem 1rem; height: 100%; }
.law-card .hd { font-size: 0.75rem; font-weight: 800; color: #1d4ed8; margin-bottom: 6px; }
.co-card { background: #f0faf4; border: 1px solid #b9e4c9; border-radius: 12px; padding: 0.9rem 1rem; height: 100%; }
.co-card .hd { font-size: 0.75rem; font-weight: 800; color: #15803d; margin-bottom: 6px; }
.item-name { font-size: 0.88rem; font-weight: 700; color: #1e293b; }
.item-desc { font-size: 0.76rem; color: #475569; line-height: 1.55; margin: 1px 0 7px; }

.script-box { background: #fff; border: 1px solid #e2e8f0; border-left: 4px solid #2563eb; border-radius: 0 10px 10px 0;
  padding: 0.8rem 1rem; font-size: 0.95rem; color: #1e3a5f; line-height: 1.7; margin-bottom: 0.5rem; }
.warn-box { background: #fef2f2; border: 1px solid #fecaca; border-radius: 8px; padding: 0.5rem 0.8rem;
  font-size: 0.78rem; font-weight: 600; color: #b91c1c; margin-top: 6px; line-height: 1.5; }
.note-box { background: #fff8ed; border: 1px solid #fbd9a5; border-radius: 8px; padding: 0.5rem 0.8rem;
  font-size: 0.78rem; color: #92400e; margin-top: 6px; line-height: 1.5; }
.form-chip { display: inline-block; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;
  padding: 4px 10px; font-size: 0.78rem; font-weight: 600; color: #334155; margin: 0 5px 5px 0; }

.faq-q { font-size: 0.82rem; font-weight: 700; color: #1d4ed8; margin-bottom: 2px; }
.faq-a { font-size: 0.79rem; color: #475569; line-height: 1.55; margin-bottom: 8px; }

.case-row { background: #fff; border: 1px solid #e2e8f0; border-radius: 10px; padding: 0.6rem 0.9rem;
  display: flex; justify-content: space-between; align-items: center; margin-bottom: 6px; font-size: 0.82rem; }
.dday { font-size: 0.74rem; font-weight: 800; padding: 2px 8px; border-radius: 10px; }

.bar-seg { height: 26px; display: flex; align-items: center; justify-content: center;
  font-size: 0.72rem; font-weight: 700; color: #1e293b; white-space: nowrap; overflow: hidden; }

.footer-note { font-size: 0.74rem; color: #94a3b8; margin-top: 1.6rem; line-height: 1.6;
  border-top: 1px solid #e2e8f0; padding-top: 0.7rem; }

.stButton button { border-radius: 8px !important; font-weight: 700 !important; font-size: 0.82rem !important; }
div[data-testid="stChatInput"] textarea { font-size: 0.88rem !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 데이터: 단계 (STEP 0~8)  ※ 개정 시 이 블록만 수정
# ─────────────────────────────────────────────────────────────
STEPS = [
    {
        "id": 0, "short": "난임 지원", "title": "난임치료휴가 (임신 준비 단계)",
        "target": "난임치료 중 직원",
        "guide": "난임치료휴가는 연간 6일까지 사용 가능하며 최초 2일은 유급입니다. 신청 사실과 사유는 철저히 비밀로 유지되니 안심하고 신청해 주세요.",
        "legal": [
            {"name": "난임치료휴가 연 6일 (유급 2일)", "desc": "2025년 개정으로 3일(유급 1일)에서 확대. 사용 사실 비밀유지 의무가 법에 명시됨.", "law": "남녀고용평등법 제18조의3"},
        ],
        "company": [],
        "check": ["비밀유지 약속 및 신청 절차 안내", "유급 2일 / 무급 4일 구분 안내", "증빙(진료 확인서 등) 최소 수준으로 안내"],
        "forms": ["난임치료 관련 진료 확인 서류"],
        "warn": ["난임치료휴가 청구·사용을 이유로 한 불리한 처우 금지", "사용 사실 누설 금지 (법정 비밀유지 의무)"],
        "faq": [
            {"q": "배우자 난임치료 동행도 사용 가능한가요?", "a": "난임치료휴가는 치료를 받는 근로자 본인에게 부여됩니다. 배우자 동행 목적은 연차 사용을 안내해 주세요."},
            {"q": "증빙으로 어떤 서류가 필요한가요?", "a": "난임치료 사실을 확인할 수 있는 서류(진료확인서 등)면 충분합니다. 치료 내용 상세는 요구하지 않습니다."},
        ],
    },
    {
        "id": 1, "short": "임신 확인", "title": "임신 확인 및 초기 응대",
        "target": "임신 확인 직원",
        "guide": "축하드립니다. 임신 소식은 본인이 원하는 범위 내에서만 공유되니 안심하세요. 단축근무 등 사용 가능한 제도부터 안내해 드릴게요.",
        "legal": [
            {"name": "모성보호 및 불리한 처우 금지", "desc": "임신을 이유로 한 해고·불이익 처우 금지.", "law": "근로기준법 제74조, 남녀고용평등법"},
        ],
        "company": [],
        "check": ["공유 희망 범위 확인 및 비밀유지 약속", "이용 가능 제도 요약 안내 (단축·검진·휴가)", "플로우 내 신청서 경로 안내"],
        "forms": ["임신확인서", "KCIM_임신·육아기 관련 지원 신청서"],
        "warn": ["본인 동의 없는 임신 사실 공유 금지", "임신을 이유로 한 업무 배치 불이익 금지"],
        "faq": [
            {"q": "직원이 임신 사실을 비밀로 해달라고 하면?", "a": "당연히 비밀 유지됩니다. 본인 동의 없이 공유하지 않음을 먼저 안심시켜 주세요."},
            {"q": "임신 초기라 단축 신청을 미루고 싶다면?", "a": "제도 존재만 안내하고, 필요할 때 언제든 신청 가능하다고 전달하면 됩니다."},
        ],
    },
    {
        "id": 2, "short": "임신기 단축", "title": "임신기 근로시간 단축·출퇴근시간 변경",
        "target": "임신 중 직원",
        "guide": "임신 12주 이내 또는 32주 이후에는 하루 2시간 단축 근무가 가능하고 급여는 그대로 유지됩니다. 그 사이 기간에는 출퇴근 시간 변경을 신청하실 수 있어요.",
        "legal": [
            {"name": "임신기 근로시간 단축 (1일 2시간)", "desc": "12주 이내 또는 32주 이후. 고위험 임산부(유산·조산 위험 진단)는 전 기간 가능. 임금 삭감 금지.", "law": "근로기준법 제74조 제7항"},
            {"name": "출퇴근시간 변경 신청", "desc": "단축 불가 기간(13~31주)에도 소정근로시간을 유지하며 출퇴근 시각 변경 신청 가능.", "law": "근로기준법 제74조 제9항·제10항"},
        ],
        "company": [],
        "check": ["임신 주수 확인 (12주 이내 / 32주 이후 / 고위험)", "단축 또는 출퇴근 변경 시간대 협의", "팀 내 업무 조정 지원"],
        "forms": ["KCIM_임신·육아기 관련 지원 신청서", "(고위험 시) 의사 소견서"],
        "warn": ["단축 기간에도 급여 전액 유지 (임금 삭감 금지)"],
        "faq": [
            {"q": "13~31주에는 아무것도 못 하나요?", "a": "단축은 불가하지만 출퇴근시간 변경을 신청할 수 있습니다. 고위험 임산부 진단 시에는 전 기간 단축 가능합니다."},
            {"q": "단축 2시간을 나눠 쓸 수 있나요?", "a": "출근 전·후로 나눠 사용 가능합니다. 본인이 선택할 수 있어요."},
        ],
    },
    {
        "id": 3, "short": "태아검진", "title": "정기 건강진단 (태아검진 시간)",
        "target": "검진 대상 직원",
        "guide": "정기 태아검진 시간은 유급으로 보장됩니다. 신청서만 작성해 주시면 됩니다.",
        "legal": [
            {"name": "태아검진 시간 유급 보장", "desc": "28주 미만 월 1회 / 28~36주 2주 1회 / 36주 이후 주 1회 기준.", "law": "근로기준법 제74조의2"},
        ],
        "company": [],
        "check": ["임신 주수별 검진 주기 확인", "증빙(예약 문자·영수증) 안내", "플로우 신청 경로 안내"],
        "forms": ["KCIM_임신·육아기 관련 지원 신청서", "검진 증빙(예약 문자 또는 영수증)"],
        "warn": ["검진 결과 자체는 개인 의료정보 — 제출 요구 금지"],
        "faq": [
            {"q": "검진 결과를 회사에 내야 하나요?", "a": "아니요. 검진 사실 증빙만 필요하며 결과는 제출 의무가 없습니다."},
        ],
    },
    {
        "id": 4, "short": "연차 정리", "title": "연차 정리 및 인수인계",
        "target": "출산휴가 예정 직원",
        "guide": "출산휴가 전에 남은 연차를 이어 붙여 조금 더 일찍 쉬실 수 있어요. 일정은 연계 플래너로 함께 잡아드릴게요.",
        "legal": [
            {"name": "연차 자율 사용 원칙", "desc": "연차는 근로자가 청구한 시기에 부여. 강제 소진 지시는 불가.", "law": "근로기준법 제60조"},
        ],
        "company": [],
        "check": ["잔여 연차 확인 및 사용 계획 협의", "연계 플래너로 전체 일정표 작성", "인수인계 항목 리스트 지원"],
        "forms": ["어울지기 내 연차 신청"],
        "warn": ["연차 강제 소진 지시 금지 (자율 사용 원칙)"],
        "note": "차년도 연차 선사용은 법정 의무가 아닌 회사 재량 사항입니다. 허용 여부·한도는 내부 규정 확정 후 안내하세요. (미확정 시 경영관리본부 협의)",
        "faq": [
            {"q": "출산휴가 전 연차를 다 써야 하나요?", "a": "아닙니다. 연차는 직원이 원하는 시점에 자유롭게 사용합니다."},
            {"q": "출산일이 예정보다 빨라지면?", "a": "출산휴가는 실제 출산일 기준으로 재산정됩니다. 즉시 담당자에게 알려달라고 안내하세요. 연계 플래너에 실제 출산일을 입력하면 자동 재계산됩니다."},
        ],
    },
    {
        "id": 5, "short": "출산", "title": "출산전후휴가·배우자 출산휴가",
        "target": "출산 전후 직원 (본인·배우자)",
        "guide": "출산휴가는 90일(다태아 120일, 미숙아 100일)이고 산후 45일 이상이 보장됩니다. 배우자분은 20일 유급 휴가를 3회까지 나눠 쓰실 수 있어요. 회사에서 경조지원금 50만 원도 지급됩니다.",
        "legal": [
            {"name": "출산전후휴가 90일", "desc": "다태아 120일(산후 60일 보장) / 미숙아 100일(2025 신설) / 일반 산후 45일 이상 보장.", "law": "근로기준법 제74조"},
            {"name": "배우자 출산휴가 20일 유급", "desc": "출산일부터 120일 이내, 3회 분할 사용 가능 (2025 개정: 10일→20일, 분할 확대).", "law": "남녀고용평등법 제18조의2"},
            {"name": "유산·사산휴가", "desc": "임신 11주 이내 10일(2025 확대) ~ 28주 이상 90일. 의사 진단서 필요.", "law": "근로기준법 제74조 제3항"},
        ],
        "company": [
            {"name": "자녀출산 경조지원금 50만 원", "desc": "본인 출산·배우자 출산 동일 적용. 제출서류: 출산확인서, 가족관계증명서.", "src": "KCIM 2026 복지제도"},
            {"name": "경조사 접수 절차", "desc": "팀장 및 경영관리본부 담당자(이경한 매니저)에게 고지 → 서류 확인 → 지급 처리.", "src": "KCIM 2026 복지제도"},
        ],
        "check": ["출산확인서·가족관계증명서 수령 확인", "경조지원금 50만 원 지급 처리 (본인/배우자 동일)", "배우자 휴가 기한(출산일+120일)·분할 이력 기록", "미숙아·다태아 여부 확인 (휴가일수 분기)"],
        "forms": ["어울지기 내 신청", "출산확인서", "가족관계증명서"],
        "warn": ["산후 45일(다태아 60일) 미보장 시 법 위반 — 일정 확정 전 플래너로 검증", "배우자 휴가는 출산일부터 120일 이내 사용"],
        "faq": [
            {"q": "배우자 출산휴가 급여는 회사 부담인가요?", "a": "20일 전체 유급입니다. 우선지원대상기업 여부에 따라 고용보험 급여 지원이 달라질 수 있으니 관할 고용센터에 확인하세요."},
            {"q": "유산한 직원이 있다면?", "a": "임신 주수에 따라 10~90일 유급 휴가가 보장됩니다(11주 이내도 2025년부터 10일). 진단서 제출 후 신청 안내해 주세요."},
        ],
    },
    {
        "id": 6, "short": "육아기 단축", "title": "육아기 근로시간 단축",
        "target": "육아기 부모 직원",
        "guide": "자녀가 만 12세 이하 또는 초등 6학년 이하라면 단축 근무가 가능합니다. 기본 1년에 육아휴직 미사용 기간의 2배를 더해 최대 3년까지 쓸 수 있어요.",
        "legal": [
            {"name": "육아기 근로시간 단축", "desc": "자녀 만 12세 이하 또는 초6 이하(2025 확대). 기간 = 기본 1년 + 육아휴직 미사용기간×2 (최대 3년). 단축 후 주 15~35시간, 최소 1개월 단위.", "law": "남녀고용평등법 제19조의2"},
        ],
        "company": [],
        "check": ["자녀 연령 확인 (만 12세 이하/초6 이하)", "육아휴직 잔여일 확인 → 단축 가능 기간 산출", "단축 시간·기간 협의 (최소 1개월)"],
        "forms": ["KCIM_임신·육아기 관련 지원 신청서"],
        "warn": ["급여는 단축 시간에 비례 (고용보험 지원금 별도 확인)", "연차는 단축 시간 비례 산정"],
        "faq": [
            {"q": "육아휴직을 안 썼다면 단축은 몇 년까지 가능한가요?", "a": "기본 1년 + 미사용 육아휴직(최대 1년)×2 = 최대 3년까지 가능합니다."},
            {"q": "육아휴직과 동시에 쓸 수 있나요?", "a": "동시 사용은 불가하며, 순차 전환(휴직→단축, 단축→휴직)은 가능합니다."},
        ],
    },
    {
        "id": 7, "short": "육아휴직·복직", "title": "육아휴직 및 복직 관리",
        "target": "휴직·복직 예정 직원",
        "guide": "육아휴직은 기본 1년이고, 부모가 각각 3개월 이상 사용하는 등 요건을 충족하면 1년 6개월까지 연장됩니다. 급여는 사후지급 없이 매월 전액 지급돼요.",
        "legal": [
            {"name": "육아휴직 1년 (요건 충족 시 1년 6개월)", "desc": "연장 요건: 부모 각각 3개월 이상 사용 / 한부모 / 중증 장애아동 부모. 3회 분할 가능(최대 4개 구간).", "law": "남녀고용평등법 제19조"},
            {"name": "육아휴직 급여 (사후지급금 폐지)", "desc": "1~3개월 월 최대 250만, 4~6개월 200만, 7개월~ 160만 원. 6+6 부모육아휴직제: 부모 각각 첫 6개월 상향 지원.", "law": "고용보험법 시행령 (2025 개정)"},
            {"name": "복직 보장", "desc": "휴직 전과 동일하거나 동등한 수준의 업무·임금 복귀 보장. 불리한 처우 금지.", "law": "남녀고용평등법 제19조 제3항·제4항"},
        ],
        "company": [
            {"name": "Refresh 휴가 (참고)", "desc": "근속 3·5·10·15년차 부여, 분할 금지, 팀장 전결. 복직 후 장기근속 시 활용 안내.", "src": "KCIM 2026 복지제도"},
        ],
        "check": ["연장 요건(부모 각 3개월 등) 해당 여부 확인 → 기간 안내", "복직 예정일 확인 및 사전 면담 일정 (권장: 복직 14일 전)", "자리·계정·업무 세팅 및 복귀 프로그램", "복직 후 처우 동일성 점검"],
        "forms": ["어울지기 내 신청 (휴직/복직)"],
        "warn": ["복직 후 부당 처우·차별 절대 금지", "복직 14일 전 면담은 사내 권장 절차 (법정 기한 아님)"],
        "faq": [
            {"q": "1년 6개월은 누구나 가능한가요?", "a": "아닙니다. 부모 각각 3개월 이상 사용, 한부모, 중증 장애아동 부모 요건 중 하나를 충족해야 연장됩니다."},
            {"q": "부부가 동시에 육아휴직이 가능한가요?", "a": "가능합니다. 6+6 부모육아휴직제로 부모 각각 첫 6개월 급여가 상향 지원됩니다."},
        ],
    },
    {
        "id": 8, "short": "가족돌봄·성장", "title": "가족돌봄 및 자녀 성장 지원",
        "target": "가족돌봄·자녀 성장기 직원",
        "guide": "가족 돌봄이 필요하면 연 90일의 가족돌봄휴직과 연 10일의 가족돌봄휴가를 쓸 수 있습니다. 자녀 대학 입학 시 회사에서 100만 원을 지원해 드려요.",
        "legal": [
            {"name": "가족돌봄휴직 연 90일", "desc": "1회 30일 이상 단위. 가족의 질병·사고·노령 사유.", "law": "남녀고용평등법 제22조의2"},
            {"name": "가족돌봄휴가 연 10일", "desc": "1일 단위 사용 가능 (무급). 자녀 돌봄 사유 포함.", "law": "남녀고용평등법 제22조의2"},
        ],
        "company": [
            {"name": "자녀 대학교 입학지원 100만 원", "desc": "1년 이상 근무자. 제출서류: 입학사실확인서(없는 경우 재학증명서).", "src": "KCIM 2026 복지제도"},
            {"name": "자녀결혼 경조지원 100만 원", "desc": "제출서류: 자녀 결혼 청첩장 사본.", "src": "KCIM 2026 복지제도"},
        ],
        "check": ["돌봄 사유·기간 확인 (휴직 30일 단위 / 휴가 1일 단위)", "자녀 성장 지원 해당 여부 확인 (입학·결혼)", "서류 수령 및 지급 처리"],
        "forms": ["가족돌봄 신청서", "(입학지원) 입학사실확인서", "(자녀결혼) 청첩장 사본"],
        "warn": ["가족돌봄 사용을 이유로 한 불리한 처우 금지"],
        "faq": [
            {"q": "가족돌봄휴가는 유급인가요?", "a": "법정 기준은 무급입니다. 회사 자체 유급 여부는 경영관리본부에 확인하세요."},
        ],
    },
]

SCENARIOS = [
    {"icon": "💜", "name": "난임·임신 준비", "desc": "난임치료휴가 안내", "step": 0},
    {"icon": "🤰", "name": "임신 확인", "desc": "초기 응대 → 단축·검진", "step": 1},
    {"icon": "👨‍👩‍👧", "name": "출산 (본인·배우자)", "desc": "휴가 + 경조지원금 50만", "step": 5},
    {"icon": "🏢", "name": "복직 준비", "desc": "면담·자리 세팅", "step": 7},
    {"icon": "🧑‍🤝‍🧑", "name": "가족돌봄·자녀 성장", "desc": "돌봄휴직·입학지원", "step": 8},
]

# ─────────────────────────────────────────────────────────────
# 유틸: 케이스 저장/불러오기, 날짜 계산
# ─────────────────────────────────────────────────────────────
def load_cases():
    if CASES_FILE.exists():
        try:
            return json.loads(CASES_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def save_cases(cases):
    try:
        CASES_FILE.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        st.warning(f"케이스 저장 실패: {e} (배포 환경에 따라 파일 저장이 제한될 수 있습니다)")

def weekend_count(s: date, e: date) -> int:
    return sum(1 for i in range((e - s).days + 1) if (s + timedelta(days=i)).weekday() >= 5)

def fmt(d: date) -> str:
    return d.strftime("%Y-%m-%d")

def get_case(cases, cid):
    for c in cases:
        if c["id"] == cid:
            return c
    return None

# ─────────────────────────────────────────────────────────────
# 세션 초기화
# ─────────────────────────────────────────────────────────────
ss = st.session_state
ss.setdefault("view", "home")           # home | step | calc
ss.setdefault("active_step", 1)
ss.setdefault("calc_tab_hint", 0)
ss.setdefault("cases", load_cases())
ss.setdefault("active_case", None)
ss.setdefault("chat_open", False)
ss.setdefault("messages", [])
ss.setdefault("pending_q", None)

def go(view, step=None, case=None):
    ss.view = view
    if step is not None:
        ss.active_step = step
    if case is not None:
        ss.active_case = case
    st.rerun()

# ─────────────────────────────────────────────────────────────
# 상단 헤더
# ─────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="top-header">
  <div>
    <div class="top-title">🤝 KCIM 일가정양립 지원 가이드</div>
    <div class="top-sub">경영관리본부 담당자용 · 난임 → 임신 → 출산 → 육아 → 돌봄 전 여정</div>
  </div>
  <div class="ver-badge">{VERSION_BASIS}</div>
</div>
""", unsafe_allow_html=True)

nav1, nav2, nav3, _sp = st.columns([1, 1, 1, 4])
with nav1:
    if st.button("🏠 홈", use_container_width=True, type="primary" if ss.view == "home" else "secondary"):
        go("home")
with nav2:
    if st.button("📖 단계 가이드", use_container_width=True, type="primary" if ss.view == "step" else "secondary"):
        go("step")
with nav3:
    if st.button("🧮 계산 도구", use_container_width=True, type="primary" if ss.view == "calc" else "secondary"):
        go("calc")

# ═════════════════════════════════════════════════════════════
# 홈 대시보드
# ═════════════════════════════════════════════════════════════
if ss.view == "home":
    # 검색
    q = st.text_input("🔍 제도·키워드 검색", placeholder="예: 배우자 출산, 단축, 경조지원금", key="home_search")
    if q:
        hits = []
        ql = q.lower()
        for i, s in enumerate(STEPS):
            blob = " ".join([s["title"], s["short"], s["guide"]]
                            + [l["name"] + l["desc"] for l in s["legal"]]
                            + [c["name"] + c["desc"] for c in s.get("company", [])]
                            + [f["q"] + f["a"] for f in s["faq"]]).lower()
            if ql in blob:
                hits.append(i)
        if hits:
            st.markdown('<div class="sec-title">검색 결과</div>', unsafe_allow_html=True)
            for i in hits:
                c1, c2 = st.columns([5, 1])
                c1.markdown(f"**STEP {STEPS[i]['id']}. {STEPS[i]['title']}** — {STEPS[i]['guide'][:60]}…")
                if c2.button("이동", key=f"srch_{i}", use_container_width=True):
                    go("step", step=i)
        else:
            st.info("검색 결과가 없습니다. 다른 키워드로 시도해 보세요.")

    # 상황별 진입
    st.markdown('<div class="sec-title">어떤 상황인가요? — 상황별 바로가기</div>', unsafe_allow_html=True)
    sc_cols = st.columns(len(SCENARIOS))
    for col, sc in zip(sc_cols, SCENARIOS):
        with col:
            st.markdown(f"""<div class="scn-card">
              <div style="font-size:1.3rem;">{sc['icon']}</div>
              <div class="scn-name">{sc['name']}</div>
              <div class="scn-desc">{sc['desc']}</div></div>""", unsafe_allow_html=True)
            if st.button("열기", key=f"scn_{sc['step']}", use_container_width=True):
                go("step", step=sc["step"])

    col_case, col_tool = st.columns([1.6, 1])

    # 진행 중 케이스
    with col_case:
        st.markdown('<div class="sec-title">📋 진행 중 케이스</div>', unsafe_allow_html=True)
        if not ss.cases:
            st.markdown('<div class="case-row" style="color:#94a3b8;">등록된 케이스가 없습니다. 아래에서 추가하세요.</div>', unsafe_allow_html=True)
        today = date.today()
        for c in ss.cases:
            badge = ""
            if c.get("return_date"):
                rd = date.fromisoformat(c["return_date"])
                dd = (rd - today).days
                if dd >= 0:
                    color = "#dc2626" if dd <= 30 else "#2563eb"
                    badge = f'<span class="dday" style="background:{color}1a;color:{color};">복직 D-{dd}</span>'
            r1, r2, r3 = st.columns([4, 1, 1])
            r1.markdown(f'<div class="case-row"><span><b>{c["name"]}</b> · STEP {c.get("current_step", 1)} '
                        f'{STEPS[c.get("current_step", 1)]["short"]} · 예정일 {c.get("due_date", "-")}</span>{badge}</div>',
                        unsafe_allow_html=True)
            if r2.button("열기", key=f"case_open_{c['id']}", use_container_width=True):
                go("step", step=c.get("current_step", 1), case=c["id"])
            if r3.button("삭제", key=f"case_del_{c['id']}", use_container_width=True):
                ss.cases = [x for x in ss.cases if x["id"] != c["id"]]
                save_cases(ss.cases)
                st.rerun()

        with st.expander("➕ 새 케이스 추가"):
            n1, n2, n3 = st.columns([2, 2, 1])
            new_name = n1.text_input("직원명", key="new_case_name", placeholder="예: 김지은 책임")
            new_due = n2.date_input("출산예정일", key="new_case_due", value=today + timedelta(days=120))
            n3.markdown("<br>", unsafe_allow_html=True)
            if n3.button("추가", key="new_case_btn", use_container_width=True):
                if new_name.strip():
                    ss.cases.append({"id": f"c{int(pd.Timestamp.now().timestamp())}", "name": new_name.strip(),
                                     "due_date": fmt(new_due), "current_step": 1, "checks": {}, "return_date": None})
                    save_cases(ss.cases)
                    st.rerun()
                else:
                    st.warning("직원명을 입력하세요.")

    # 빠른 도구
    with col_tool:
        st.markdown('<div class="sec-title">🧮 빠른 도구</div>', unsafe_allow_html=True)
        if st.button("📆 출산휴가 연계 플래너 (연차→출산→육아휴직)", use_container_width=True):
            ss.calc_tab_hint = 0
            go("calc")
        if st.button("📐 임신기 단축 기간 계산", use_container_width=True):
            ss.calc_tab_hint = 1
            go("calc")
        if st.button("📉 무급휴가 연차 삭감 계산", use_container_width=True):
            ss.calc_tab_hint = 2
            go("calc")
        st.markdown("""<div class="note-box" style="margin-top:10px;">
          경조지원금·복지 문의: 경영관리본부 (이경한 매니저)<br>
          기타 복지제도는 별도 공지 '2026 복지제도'를 참고하세요.</div>""", unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════
# 단계 가이드
# ═════════════════════════════════════════════════════════════
elif ss.view == "step":
    idx = ss.active_step
    step = STEPS[idx]

    # 케이스 선택
    case = get_case(ss.cases, ss.active_case) if ss.active_case else None
    case_names = ["(케이스 없이 보기)"] + [f'{c["name"]} (예정일 {c["due_date"]})' for c in ss.cases]
    cur_sel = 0
    if case:
        cur_sel = 1 + [c["id"] for c in ss.cases].index(case["id"])
    sel = st.selectbox("적용할 케이스", case_names, index=cur_sel, key="case_selector")
    new_case = None if sel == case_names[0] else ss.cases[case_names.index(sel) - 1]
    if (new_case["id"] if new_case else None) != ss.active_case:
        ss.active_case = new_case["id"] if new_case else None
        st.rerun()
    case = new_case

    done_set = set()
    if case:
        for k, v in case.get("checks", {}).items():
            if v and all(v):
                done_set.add(int(k))

    # 여정 타임라인
    tl = '<div class="tl-wrap">'
    for i, s in enumerate(STEPS):
        cls = "now" if i == idx else ("done" if i in done_set else "")
        mark = "✓" if (i in done_set and i != idx) else str(s["id"])
        tl += f'<div class="tl-dot {cls}">{mark}</div>'
        if i < len(STEPS) - 1:
            tl += '<div class="tl-line"></div>'
    tl += f'<div class="tl-label">STEP {step["id"]} · {step["short"]}</div></div>'
    st.markdown(tl, unsafe_allow_html=True)

    # 단계 이동 버튼
    mv = st.columns(len(STEPS))
    for i, s in enumerate(STEPS):
        if mv[i].button(f'{s["id"]}. {s["short"]}', key=f"mv_{i}", use_container_width=True,
                        type="primary" if i == idx else "secondary"):
            go("step", step=i)

    # 헤더
    case_tag = f' · 케이스: {case["name"]}' if case else ""
    st.markdown(f"""<div class="step-head">
      <span class="num">STEP {step['id']} / {len(STEPS)-1}</span>
      <div class="ttl">{step['title']}</div>
      <div class="meta">대상: {step['target']}{case_tag}</div></div>""", unsafe_allow_html=True)

    # 법정 / 회사 2단 카드
    cl, cr = st.columns(2)
    with cl:
        html = '<div class="law-card"><div class="hd">⚖️ 법정 제도</div>'
        for l in step["legal"]:
            html += f'<div class="item-name">{l["name"]}</div><div class="item-desc">{l["desc"]}<br>근거: {l["law"]}</div>'
        html += "</div>"
        st.markdown(html, unsafe_allow_html=True)
    with cr:
        html = '<div class="co-card"><div class="hd">🏢 KCIM 회사 지원</div>'
        if step.get("company"):
            for c_it in step["company"]:
                html += f'<div class="item-name">{c_it["name"]}</div><div class="item-desc">{c_it["desc"]}<br>출처: {c_it["src"]}</div>'
        else:
            html += '<div class="item-desc">이 단계에 해당하는 회사 자체 지원 항목은 없습니다.<br>전체 복지제도는 별도 공지를 참고하거나 경영관리본부로 문의하세요.</div>'
        html += "</div>"
        st.markdown(html, unsafe_allow_html=True)

    st.markdown("")

    # 스크립트 + 체크리스트
    left, right = st.columns([1.4, 1])
    with left:
        st.markdown('<div class="sec-title">💬 담당자 안내 스크립트</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="script-box">"{step["guide"]}"</div>', unsafe_allow_html=True)
        st.code(step["guide"], language=None)  # 복사 버튼용

        st.markdown('<div class="sec-title">✅ 담당자 체크리스트</div>', unsafe_allow_html=True)
        key_prefix = f'chk_{case["id"] if case else "global"}_{idx}'
        saved = (case or {}).get("checks", {}).get(str(idx), [False] * len(step["check"])) if case \
            else ss.get(f"g_checks_{idx}", [False] * len(step["check"]))
        if len(saved) != len(step["check"]):
            saved = [False] * len(step["check"])
        new_vals = []
        for ci, txt in enumerate(step["check"]):
            new_vals.append(st.checkbox(txt, value=saved[ci], key=f"{key_prefix}_{ci}"))
        if case:
            case.setdefault("checks", {})[str(idx)] = new_vals
            case["current_step"] = idx
            save_cases(ss.cases)
        else:
            ss[f"g_checks_{idx}"] = new_vals
        st.progress(sum(new_vals) / len(new_vals))

    with right:
        st.markdown('<div class="sec-title">🧾 필요 서류</div>', unsafe_allow_html=True)
        st.markdown("".join(f'<span class="form-chip">📄 {f}</span>' for f in step["forms"]), unsafe_allow_html=True)
        st.markdown('<div class="sec-title">⚠️ 주의사항</div>', unsafe_allow_html=True)
        for w in step["warn"]:
            st.markdown(f'<div class="warn-box">{w}</div>', unsafe_allow_html=True)
        if step.get("note"):
            st.markdown(f'<div class="note-box">📌 {step["note"]}</div>', unsafe_allow_html=True)
        if idx == 5 or idx == 4:
            if st.button("📆 연계 플래너로 일정 잡기", use_container_width=True):
                ss.calc_tab_hint = 0
                go("calc")

    # FAQ
    st.markdown('<div class="sec-title">💡 자주 묻는 질문</div>', unsafe_allow_html=True)
    faq_html = ""
    for f in step["faq"]:
        faq_html += f'<div class="faq-q">Q. {f["q"]}</div><div class="faq-a">A. {f["a"]}</div>'
    st.markdown(faq_html, unsafe_allow_html=True)

    # ── AI 상담 (접이식) ──
    st.markdown("---")
    if st.button(("🔽 AI 상담 닫기" if ss.chat_open else "💬 AI 상담 열기 (육아지원박사)"), key="chat_toggle"):
        ss.chat_open = not ss.chat_open
        st.rerun()

    if ss.chat_open:
        st.caption("⚠️ 외부 AI API(OpenAI)로 전송됩니다. 직원 실명 등 개인정보·민감정보는 입력하지 마세요. (사내 보안정책 검토 전)")
        pc = st.columns(3)
        presets = ["이 단계에서 가장 자주 하는 실수는?", "직원에게 보낼 안내 문자 초안 작성해줘", "관련 법 조항을 요약해줘"]
        for pcol, p in zip(pc, presets):
            if pcol.button(p, key=f"preset_{p[:6]}", use_container_width=True):
                ss.pending_q = p
                st.rerun()

        for msg in ss.messages:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        user_q = st.chat_input("질문을 입력하세요 (개인정보 제외)")
        if ss.pending_q and not user_q:
            user_q = ss.pending_q
            ss.pending_q = None

        if user_q:
            ss.messages.append({"role": "user", "content": user_q})
            with st.chat_message("user"):
                st.markdown(user_q)
            try:
                from openai import OpenAI
                if "OPENAI_API_KEY" not in st.secrets:
                    st.error("OPENAI_API_KEY가 secrets.toml에 없습니다.")
                    st.stop()
                client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
                company_ctx = "\n".join(
                    f"- {c['name']}: {c['desc']}" for s in STEPS for c in s.get("company", []))
                sys_prompt = f"""너는 '육아지원박사', KCIM 경영관리본부 HR 담당자를 돕는 모성보호 전문 AI다.

[현재 단계] STEP {step['id']} {step['title']}
[단계 법정 제도] {[l['name'] + ' (' + l['law'] + ')' for l in step['legal']]}
[KCIM 회사 지원 규정 (이것만 회사 규정의 근거로 사용)]
{company_ctx}

[답변 원칙 — 반드시 준수]
1. 2025-02-23 시행 개정법 기준으로 답변하고, 법령 조항명을 명시할 것.
2. 확실하지 않은 내용은 지어내지 말고 "고용노동부/노무사 확인 필요"라고 답할 것.
3. KCIM 내부 규정은 위에 제공된 것만 근거로 사용하고, 제공되지 않은 사내 규정 질문(예: 휴직자 복지 적용 범위)은 "경영관리본부 확인 필요"로 안내할 것.
4. 담당자가 직원에게 바로 말할 수 있는 구어체 문장을 포함할 것.
5. 답변은 3~5문장 핵심 위주로 간결하게."""
                with st.chat_message("assistant"):
                    ph = st.empty()
                    full = ""
                    stream = client.chat.completions.create(
                        model="gpt-4o",
                        messages=[{"role": "system", "content": sys_prompt}, *ss.messages],
                        stream=True,
                    )
                    for chunk in stream:
                        full += chunk.choices[0].delta.content or ""
                        ph.markdown(full + "▌")
                    ph.markdown(full)
                ss.messages.append({"role": "assistant", "content": full})
            except Exception as e:
                st.error(f"챗봇 오류: {e}")

# ═════════════════════════════════════════════════════════════
# 계산 도구
# ═════════════════════════════════════════════════════════════
elif ss.view == "calc":
    tab_names = ["📆 출산휴가 연계 플래너", "📐 임신기 단축 기간", "📉 무급휴가 연차 삭감"]
    t1, t2, t3 = st.tabs(tab_names)

    # ── Tab 1: 출산휴가 연계 플래너 ──
    with t1:
        st.markdown("#### 출산휴가 연계 플래너 (연차 → 출산휴가 → 리프레시 → 육아휴직 → 복직)")
        st.caption("예정일과 연차 시작일만 입력하면 전체 일정이 자동 연결됩니다. 주말은 자동 계산되며, 공휴일은 제외 일수에 수기 입력하세요.")

        i1, i2, i3 = st.columns(3)
        with i1:
            emp = st.text_input("직원명", placeholder="예: 김지은 책임", key="pl_emp")
            due = st.date_input("출산예정일", value=date.today() + timedelta(days=60), key="pl_due")
            actual = st.date_input("실제 출산일 (출산 후 입력 — 자동 재산정)", value=None, key="pl_actual")
        with i2:
            use_annual = st.checkbox("출산휴가 전 연차 사용", value=True, key="pl_use_annual")
            annual_start = st.date_input("연차 시작일", value=date.today() + timedelta(days=10), key="pl_astart",
                                         disabled=not use_annual)
            holiday_excl = st.number_input("연차 기간 내 공휴일 수 (제외)", 0, 30, 0, key="pl_holi",
                                           help="주말은 자동 제외됩니다. 공휴일만 입력하세요.")
            mat_start = st.date_input("출산휴가 시작일", value=due - timedelta(days=29), key="pl_mstart",
                                      help="산후 45일(다태아 60일) 보장을 위해 산전 사용 일수에 주의하세요.")
        with i3:
            birth_type = st.selectbox("출산 유형", ["일반 (90일 / 산후 45일)", "미숙아 (100일 / 산후 45일)", "다태아 (120일 / 산후 60일)"], key="pl_type")
            refresh_days = st.number_input("리프레시 휴가 (일)", 0, 15, 0, key="pl_refresh",
                                           help="근속 3·5·10·15년차 부여. 휴직 직전 사용 가능 여부는 내부 해석 확정 필요.")
            extend = st.checkbox("육아휴직 연장 요건 충족 (부모 각 3개월 사용 / 한부모 / 중증장애아)", key="pl_extend")
            parental_limit = 548 if extend else 365
            parental_days = st.number_input("육아휴직 사용 일수", 30, parental_limit, parental_limit, key="pl_pdays",
                                            help=f"최대 {parental_limit}일. 일부만 사용하면 잔여일이 표시됩니다.")

        if st.button("🔍 일정 자동 생성", type="primary", key="pl_btn"):
            ss["plan_run"] = True
        if ss.get("plan_run"):
            dur, post_req = {"일반 (90일 / 산후 45일)": (90, 45),
                             "미숙아 (100일 / 산후 45일)": (100, 45),
                             "다태아 (120일 / 산후 60일)": (120, 60)}[birth_type]
            birth = actual or due
            rows, warns = [], []

            # 연차 (종료일 = 출산휴가 시작 전날로 자동 연결)
            if use_annual:
                a_end = mat_start - timedelta(days=1)
                if a_end < annual_start:
                    warns.append("연차 시작일이 출산휴가 시작일보다 늦습니다. 연차 구간이 계산에서 제외되었습니다.")
                else:
                    a_total = (a_end - annual_start).days + 1
                    a_wk = weekend_count(annual_start, a_end)
                    a_net = a_total - a_wk - holiday_excl
                    rows.append(["연차", fmt(annual_start), fmt(a_end), a_total, a_wk + holiday_excl, a_net])

            # 출산휴가
            mat_end = mat_start + timedelta(days=dur - 1)
            rows.append([f"출산휴가 ({dur}일)", fmt(mat_start), fmt(mat_end), dur, 0, dur])

            # 산후 검증 (기준 고정: 출산 다음날부터 종료일까지, 양끝 포함)
            post_days = (mat_end - birth).days
            rows.append(["└ 산후 기간 (검증)", fmt(birth + timedelta(days=1)), fmt(mat_end), post_days, 0, post_days])
            if birth > mat_end:
                warns.append("출산(예정)일이 출산휴가 종료일 이후입니다. 시작일을 다시 확인하세요.")
            elif post_days < post_req:
                warns.append(f"산후 기간 {post_days}일 < 법정 보장 {post_req}일 — 출산휴가 시작일을 늦추거나 실제 출산일 기준으로 재조정 필요 (근로기준법 제74조 위반 위험).")

            cursor = mat_end
            # 리프레시
            if refresh_days > 0:
                r_start = cursor + timedelta(days=1)
                r_end = r_start + timedelta(days=refresh_days - 1)
                rows.append(["리프레시", fmt(r_start), fmt(r_end), refresh_days, 0, refresh_days])
                cursor = r_end
                warns.append("리프레시 휴가의 휴직 직전 사용 가능 여부는 내부 규정 해석 확정 전입니다. 경영관리본부 확인 후 확정하세요.")

            # 육아휴직
            p_start = cursor + timedelta(days=1)
            p_end = p_start + timedelta(days=parental_days - 1)
            rows.append([f"육아휴직 ({'1.5년' if extend else '1년'})", fmt(p_start), fmt(p_end), parental_days, 0, parental_days])

            return_date = p_end + timedelta(days=1)
            interview_date = return_date - timedelta(days=14)

            df = pd.DataFrame(rows, columns=["구분", "시작일", "종료일", "총 일수", "제외(주말·공휴일)", "소계"])
            st.dataframe(df, use_container_width=True, hide_index=True)

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("복직 예정일", fmt(return_date))
            m2.metric("복직 면담 권장일 (D-14)", fmt(interview_date))
            m3.metric("산후 기간", f"{post_days}일 / 기준 {post_req}일")
            m4.metric("잔여 육아휴직", f"{max(0, parental_limit - parental_days)}일 / 한도 {parental_limit}일")

            for w in warns:
                st.warning(w)
            if not any("산후" in w for w in warns):
                st.success(f"✅ 산후 {post_days}일 ≥ 법정 {post_req}일 충족")

            st.markdown("""<div class="note-box">📌 계산 기준(사양 고정): 모든 기간은 양끝 포함, 산후 기간은 출산일 다음날부터 카운트합니다.
            출산일 당일의 산전/산후 귀속은 행정해석 확인이 필요한 영역으로, 실제 출산 시 ±1일 여유를 두고 일정을 확정하세요.
            배우자 출산휴가(20일)는 출산일부터 120일 이내 별도 관리하세요.</div>""", unsafe_allow_html=True)

            # 케이스 저장
            if emp.strip():
                if st.button("💾 이 일정을 케이스로 저장", key="pl_save"):
                    existing = next((c for c in ss.cases if c["name"] == emp.strip()), None)
                    if existing:
                        existing.update({"due_date": fmt(due), "return_date": fmt(return_date)})
                    else:
                        ss.cases.append({"id": f"c{int(pd.Timestamp.now().timestamp())}", "name": emp.strip(),
                                         "due_date": fmt(due), "current_step": 4, "checks": {},
                                         "return_date": fmt(return_date)})
                    save_cases(ss.cases)
                    st.success("케이스에 저장되었습니다. 홈 대시보드에서 복직 D-day가 표시됩니다.")

            # Excel 다운로드 (기존 업무 양식 재현)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()
                ws = wb.active
                ws.title = "출산육아일정"
                thin = Border(*[Side(style="thin", color="AAAAAA")] * 4)
                fills = {"연차": "DDEBF7", "출산휴가": "E2EFDA", "└": "FCE4EC", "리프레시": "FDE9EF", "육아휴직": "FFF2CC"}

                ws["A1"] = "■ 출산휴가 및 육아휴직 계산"
                ws["A1"].font = Font(bold=True, size=13)
                ws["D1"] = emp or "-"
                ws["E1"] = "예정일"
                ws["F1"] = fmt(due)
                headers = ["구분", "시작일", "종료일", "일수", "제외(주말·공휴일)", "소계"]
                for j, h in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=j, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")
                for r, row in enumerate(rows, start=4):
                    fill_key = next((k for k in fills if row[0].startswith(k)), None)
                    for j, v in enumerate(row, start=1):
                        cell = ws.cell(row=r, column=j, value=v)
                        cell.border = thin
                        cell.alignment = Alignment(horizontal="center")
                        if fill_key:
                            cell.fill = PatternFill("solid", fgColor=fills[fill_key])
                # 우측 요약표
                ws["H3"] = "■ 휴가별 일수계산"
                ws["H3"].font = Font(bold=True)
                summary = [(row[0], row[5]) for row in rows] + [("복직 예정일", fmt(return_date))]
                for r, (k, v) in enumerate(summary, start=4):
                    ws.cell(row=r, column=8, value=k).border = thin
                    ws.cell(row=r, column=9, value=v).border = thin
                for col, w_ in zip("ABCDEFGHI", [18, 12, 12, 8, 16, 8, 2, 16, 12]):
                    ws.column_dimensions[col].width = w_

                buf = BytesIO()
                wb.save(buf)
                st.download_button("⬇️ Excel 다운로드 (기존 양식)", data=buf.getvalue(),
                                   file_name=f"출산육아일정_{emp or '직원'}_{fmt(due)}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except ImportError:
                st.info("Excel 다운로드에는 openpyxl 패키지가 필요합니다. (pip install openpyxl)")

    # ── Tab 2: 임신기 단축 기간 ──
    with t2:
        st.markdown("#### 임신기 근로시간 단축 대상기간 계산기")
        st.caption("출산예정일 기준 추정 계산입니다. 병원 임신확인서의 주수와 다를 수 있으니 확인서 주수를 우선하세요.")
        due2 = st.date_input("출산예정일", value=date.today() + timedelta(days=180), key="sh_due")
        if st.button("🔍 단축 기간 계산", key="sh_btn", type="primary"):
            lmp = due2 - timedelta(days=279)
            week12_end = lmp + timedelta(days=83)
            week32_start = lmp + timedelta(days=217)
            r = pd.DataFrame([
                ["✅ 임신 12주 이내 (단축 가능)", fmt(lmp), fmt(week12_end), (week12_end - lmp).days + 1],
                ["⏸ 13~31주 (단축 불가 · 출퇴근시간 변경 가능)", fmt(week12_end + timedelta(1)), fmt(week32_start - timedelta(1)),
                 (week32_start - week12_end).days - 1],
                ["✅ 임신 32주 이후 (단축 가능)", fmt(week32_start), fmt(due2), (due2 - week32_start).days + 1],
            ], columns=["구간", "시작", "종료", "일수"])
            st.dataframe(r, use_container_width=True, hide_index=True)
            st.markdown("""<div class="note-box">※ 고위험 임산부(유산·조산 위험 진단)는 전 기간 단축 가능 — 의사 소견서 필요 (근로기준법 제74조 제7항).
            13~31주 구간에는 출퇴근시간 변경 신청을 안내하세요 (동조 제9항·제10항).</div>""", unsafe_allow_html=True)

    # ── Tab 3: 무급휴가 연차 삭감 ──
    with t3:
        st.markdown("#### 무급휴가 연차 삭감 계산기")
        st.caption("연간 소정근로일 80% 미만 출근 시 연차가 비례 산정됩니다 (근로기준법 제60조). 출산전후휴가·육아휴직 기간은 출근으로 간주되어 자동 제외됩니다 (동조 제6항).")

        a1, a2 = st.columns(2)
        annual_cnt = a1.number_input("발생 연차 (일)", 1, 25, 15, key="cut_annual")
        base_days = a2.number_input("연도 기준일수", 365, 366, 365, key="cut_base", help="윤년은 366")

        df_leave = st.data_editor(
            pd.DataFrame({"구분": ["개인 무급휴가"], "시작일": [date.today()], "종료일": [date.today() + timedelta(days=30)]}),
            column_config={
                "구분": st.column_config.SelectboxColumn("구분", options=[
                    "개인 무급휴가", "가족돌봄휴직(무급)", "출산전후휴가(출근간주·제외)", "육아휴직(출근간주·제외)"], required=True),
                "시작일": st.column_config.DateColumn("시작일", required=True),
                "종료일": st.column_config.DateColumn("종료일", required=True),
            },
            num_rows="dynamic", use_container_width=True, key="cut_df")

        if st.button("🔍 연차 삭감 계산", key="cut_btn", type="primary"):
            total_unpaid, excluded = 0, 0
            for _, row in df_leave.iterrows():
                s_, e_ = row["시작일"], row["종료일"]
                if s_ is None or e_ is None:
                    continue
                if hasattr(s_, "date"):
                    s_ = s_.date()
                if hasattr(e_, "date"):
                    e_ = e_.date()
                d = (e_ - s_).days + 1
                if "출근간주" in str(row["구분"]):
                    excluded += d
                else:
                    total_unpaid += d
            ratio = (base_days - total_unpaid) / base_days
            adjusted = max(0, math.ceil(annual_cnt * ratio * 2) / 2)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("무급 산입 일수", f"{total_unpaid}일")
            c2.metric("출근간주 제외", f"{excluded}일")
            c3.metric("근무 비율", f"{ratio*100:.1f}%")
            c4.metric("조정 후 연차", f"{adjusted}일 (−{annual_cnt - adjusted}일)")
            if ratio >= 0.8:
                st.success(f"근무비율 {ratio*100:.1f}% ≥ 80% → 연차 삭감 없음 ({annual_cnt}일 전액)")
            else:
                st.warning(f"근무비율 {ratio*100:.1f}% < 80% → {annual_cnt}일 → {adjusted}일로 조정 (0.5일 단위 올림)")
            st.markdown("""<div class="note-box">※ 출산전후휴가·육아휴직은 법상 출근으로 간주되어 계산에서 자동 제외했습니다 (근로기준법 제60조 제6항).
            개별 사안의 최종 판단은 노무사 확인을 권장합니다.</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 공통 푸터
# ─────────────────────────────────────────────────────────────
st.markdown(f"""<div class="footer-note">
{VERSION_BASIS}<br>
법령: 근로기준법·남녀고용평등법 (국가법령정보센터) / 급여: 고용노동부 고시 — 배포 전 최신 고시 교차 확인 권장.<br>
회사 규정 해석·미확정 사안(휴직자 복지 적용, 차년도 연차 선사용, 리프레시 사용 시점)은 경영관리본부 확인 후 안내하세요. 법률 사항의 최종 판단은 노무사 상담을 권장합니다.
</div>""", unsafe_allow_html=True)
